#!/usr/bin/env python3
"""Map historical Setaria and sorghum HSFs to the release-63 catalog."""

from __future__ import annotations

import argparse
import csv
import re
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

from select_representative_isoforms import Protein, parse_fasta


VALIDATION_FIELDS = [
    "species", "historical_id", "historical_sequence_length", "mapped_gene_id",
    "current_protein_id", "current_protein_length", "recovered_by_hmmer",
    "confirmed_by_interproscan", "sequence_status", "discrepancy_category",
    "notes",
]


def parse_alignment(path: Path) -> dict[str, str]:
    sequences: dict[str, str] = {}
    header: str | None = None
    parts: list[str] = []
    with path.open(encoding="utf-8") as handle:
        for raw_line in handle:
            line = raw_line.strip()
            if not line:
                continue
            if line.startswith(">"):
                if header is not None:
                    sequences[header] = "".join(parts).replace("-", "").replace(".", "").upper()
                header, parts = line[1:], []
            else:
                parts.append(line)
    if header is not None:
        sequences[header] = "".join(parts).replace("-", "").replace(".", "").upper()
    return sequences


def historical_ids(workbook: Path, sheet_name: str) -> list[str]:
    """Use PF00447 rows, avoiding malformed overflow rows in the old workbook."""
    book = load_workbook(workbook, read_only=True, data_only=True)
    sheet = book[sheet_name]
    identifiers = []
    seen = set()
    for row in sheet.iter_rows(min_row=2, values_only=True):
        gene, signature = row[0], row[2]
        if signature == "PF00447" and isinstance(gene, str) and gene not in seen:
            identifiers.append(gene)
            seen.add(gene)
    book.close()
    return identifiers


def current_gene_id(species: str, historical_id: str) -> str | None:
    if species == "setaria_viridis":
        match = re.match(r"^Sevir\.(\d+G\d+)\.", historical_id)
        return f"SEVIR_{match.group(1)}v2" if match else None
    if species == "sorghum_bicolor":
        match = re.match(r"^Sobic\.(\d+G\d+)\.", historical_id)
        return f"SORBI_3{match.group(1)}" if match else None
    raise ValueError(f"Unsupported historical mapping for {species}")


def read_tsv(path: Path) -> list[dict[str, str]]:
    with path.open(encoding="utf-8") as handle:
        return list(csv.DictReader(handle, delimiter="\t"))


def validate_species(
    species: str,
    sheet: str,
    workbook: Path,
    aligned_sequences: dict[str, str],
    proteome: Path,
    mapping_path: Path,
    catalog_rows: list[dict[str, str]],
) -> tuple[list[dict[str, str]], list[dict[str, str]]]:
    proteins_by_gene: dict[str, list[Protein]] = defaultdict(list)
    for protein in parse_fasta(proteome):
        proteins_by_gene[protein.gene_id].append(protein)
    selected_by_gene = {row["gene_id"]: row for row in read_tsv(mapping_path)}
    # The catalog contains both pilot species. Restrict this validation to the
    # species currently being evaluated; otherwise candidates from the other
    # species are incorrectly reported as new loci here.
    hsf_catalog = {
        row["gene_id"]: row
        for row in catalog_rows
        if row["species"] == species and row["family"] == "HSF"
    }

    rows = []
    historical_genes = set()
    for historical_id in historical_ids(workbook, sheet):
        gene_id = current_gene_id(species, historical_id)
        sequence = aligned_sequences.get(historical_id, "")
        if gene_id:
            historical_genes.add(gene_id)
        selected = selected_by_gene.get(gene_id or "")
        candidates = proteins_by_gene.get(gene_id or "", [])
        representative = next(
            (p for p in candidates if selected and p.protein_id == selected["selected_protein_id"]),
            None,
        )
        exact = [p for p in candidates if sequence and p.sequence == sequence]
        recovered = gene_id in hsf_catalog
        confirmed = recovered and hsf_catalog[gene_id]["classification_status"] in {"confirmed", "multi_family"}

        if not gene_id or gene_id not in proteins_by_gene:
            sequence_status = "not_tested"
            category = "absent_from_release_63_proteome"
            notes = "Historical locus could not be mapped to a release-63 gene."
        elif not recovered:
            sequence_status = "identical_current_isoform" if exact else "sequence_changed"
            category = "not_recovered_by_hmm_search"
            notes = "Release-63 gene exists but is absent from the HSF HMMER catalog."
        elif representative and representative in exact:
            sequence_status = "identical_to_representative"
            category = "updated_identifier"
            notes = "Historical sequence is identical to the selected release-63 protein."
        elif exact:
            sequence_status = "identical_to_alternative_isoform"
            category = "different_representative_isoform"
            notes = "Historical sequence matches a non-representative release-63 isoform."
        else:
            sequence_status = "sequence_changed"
            category = "sequence_or_annotation_inconsistency"
            notes = "Locus is recovered as HSF, but no release-63 isoform is sequence-identical."

        rows.append(
            {
                "species": species,
                "historical_id": historical_id,
                "historical_sequence_length": str(len(sequence)) if sequence else "",
                "mapped_gene_id": gene_id or "",
                "current_protein_id": selected["selected_protein_id"] if selected else "",
                "current_protein_length": selected["protein_length"] if selected else "",
                "recovered_by_hmmer": str(recovered).lower(),
                "confirmed_by_interproscan": str(confirmed).lower(),
                "sequence_status": sequence_status,
                "discrepancy_category": category,
                "notes": notes,
            }
        )

    new_rows = [
        {
            "species": species,
            "gene_id": gene_id,
            "protein_id": row["protein_id"],
            "classification_status": row["classification_status"],
            "reason": "release-63 HSF candidate absent from the historical PF00447 set",
        }
        for gene_id, row in sorted(hsf_catalog.items())
        if gene_id not in historical_genes
    ]
    return rows, new_rows


def write_tsv(path: Path, rows: list[dict[str, str]], fields: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, delimiter="\t")
        writer.writeheader()
        writer.writerows(rows)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", type=Path, default=Path("reference/external/InterproScan.xlsx"))
    parser.add_argument("--alignment", type=Path, default=Path("reference/external/Alinhamento_HSF.fas"))
    parser.add_argument("--catalog", type=Path, default=Path("results/catalog/hsp_hsf_catalog.tsv"))
    parser.add_argument("--reference-dir", type=Path, default=Path("data/reference/ensembl_plants_63"))
    parser.add_argument("--mapping-dir", type=Path, default=Path("work/isoform_mapping"))
    parser.add_argument("--outdir", type=Path, default=Path("results/catalog"))
    args = parser.parse_args()

    catalog = read_tsv(args.catalog)
    sequences = parse_alignment(args.alignment)
    settings = [
        ("setaria_viridis", "Sviridis"),
        ("sorghum_bicolor", "Sbicolor"),
    ]
    validations: list[dict[str, str]] = []
    new_candidates: list[dict[str, str]] = []
    for species, sheet in settings:
        mapped, new = validate_species(
            species,
            sheet,
            args.workbook,
            sequences,
            args.reference_dir / species / "proteins_all.fa.gz",
            args.mapping_dir / f"{species}.tsv",
            catalog,
        )
        validations.extend(mapped)
        new_candidates.extend(new)

    write_tsv(args.outdir / "historical_hsf_validation.tsv", validations, VALIDATION_FIELDS)
    new_fields = ["species", "gene_id", "protein_id", "classification_status", "reason"]
    write_tsv(args.outdir / "new_hsf_candidates.tsv", new_candidates, new_fields)
    print(
        f"Validated {len(validations)} historical HSFs; "
        f"found {len(new_candidates)} candidate additions"
    )


if __name__ == "__main__":
    main()
